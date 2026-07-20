"""Render laporan ke XLSX (sheet Rekap + Summary), gaya mirip LAPORAN/."""
from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.report import ReportResult
from app.exporters.layout import COLUMNS, HEADERS, row_values

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
_TITLE_FONT = Font(bold=True, size=14)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(vertical="top", wrap_text=True)
_STATUS_FILL = {
    "Done": "C6EFCE", "Progress": "FFEB9C", "Hold": "FFC7CE",
    "Back Log": "D9D9D9", "To Do": "BDD7EE",
}


def render_xlsx(result: ReportResult) -> bytes:
    wb = Workbook()
    _sheet_rekap(wb.active, result)
    _sheet_summary(wb.create_sheet("Summary"), result)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sheet_rekap(ws, result: ReportResult):
    ws.title = "Rekap"
    ncol = len(COLUMNS)

    ws.cell(1, 1, f"FAST REPORT — {result.source_name}").font = _TITLE_FONT
    p = result.period
    scope = {"daily": "Harian", "weekly": "Mingguan",
             "monthly": "Bulanan"}.get(p.mode, p.mode)
    ws.cell(2, 1, f"Laporan {scope}: {p.label}  (acuan: {p.date_field})")
    ws.cell(3, 1, f"Total: {len(result.rows)} task   ·   "
                  f"Dibuat: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncol)

    hrow = 5
    for c, (header, width, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(hrow, c, header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(c)].width = width

    status_col = HEADERS.index("Status") + 1
    for i, rec in enumerate(result.rows):
        r = hrow + 1 + i
        for c, value in enumerate(row_values(rec), start=1):
            cell = ws.cell(r, c, value)
            cell.alignment = _WRAP
            cell.border = _BORDER
        fill = _STATUS_FILL.get(rec.status)
        if fill:
            ws.cell(r, status_col).fill = PatternFill("solid", fgColor=fill)

    ws.freeze_panes = ws.cell(hrow + 1, 1)
    last = hrow + len(result.rows)
    ws.auto_filter.ref = f"A{hrow}:{get_column_letter(ncol)}{max(last, hrow)}"


def _sheet_summary(ws, result: ReportResult):
    s = result.summary
    ws.cell(1, 1, "RINGKASAN").font = _TITLE_FONT
    r = 3
    for title, data in (
        ("Per Status", s["by_status"]),
        ("Per Aplikasi", s["by_aplikasi"]),
        ("Per PIC", s["by_pic"]),
    ):
        ws.cell(r, 1, title).font = Font(bold=True)
        r += 1
        for k, v in data.items():
            ws.cell(r, 1, k)
            ws.cell(r, 2, v)
            r += 1
        r += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 10
