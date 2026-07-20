"""Endpoint modul Project Management (khusus superadmin).

Semua endpoint diproteksi menu 'project_mgmt' sehingga hanya role yang memiliki
menu tersebut (superadmin) yang boleh mengakses. Sumber data: sheet
"Project List 2026" (lihat app.core.project_mgmt).
"""
from __future__ import annotations

import io

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import HTMLResponse, StreamingResponse

from app.core import project_mgmt as pm
from app.security import require_menu

router = APIRouter(prefix="/api/projects")
require_pm = require_menu("project_mgmt")
require_eval = require_menu("evaluasi_tw")

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@router.get("")
def get_projects(
    section: str = "", status: str = "", pic: str = "",
    bendera: str = "", triwulan: str = "", q: str = "",
    _=Depends(require_pm),
):
    try:
        return pm.list_projects(
            section=section, status=status, pic=pic,
            bendera=bendera, triwulan=triwulan, q=q,
        )
    except FileNotFoundError as e:
        raise HTTPException(404, str(e))
    except ValueError as e:
        raise HTTPException(422, str(e))


@router.get("/filters")
def get_filters(_=Depends(require_pm)):
    try:
        return pm.filters()
    except FileNotFoundError as e:
        raise HTTPException(404, str(e))


@router.get("/summary")
def get_summary(_=Depends(require_pm)):
    try:
        return pm.summary()
    except FileNotFoundError as e:
        raise HTTPException(404, str(e))
    except ValueError as e:
        raise HTTPException(422, str(e))


@router.get("/load-mapping")
def get_load_mapping(_=Depends(require_pm)):
    try:
        return pm.load_mapping()
    except FileNotFoundError as e:
        raise HTTPException(404, str(e))


@router.get("/targets")
def get_targets(_=Depends(require_pm)):
    try:
        return pm.targets()
    except FileNotFoundError as e:
        raise HTTPException(404, str(e))


@router.get("/tw")
def get_tw(_=Depends(require_pm)):
    try:
        return pm.tw()
    except FileNotFoundError as e:
        raise HTTPException(404, str(e))


@router.post("/refresh")
def refresh(_=Depends(require_pm)):
    """Muat ulang seluruh data dari file Excel (invalidasi cache)."""
    pm.refresh()
    return {"status": "ok", "generated_at": pm.summary()["generated_at"]}


# --------------------------------------------------------------------------- #
# Evaluasi Triwulan (dokumen HTML per PIC, di-embed di frontend)
# --------------------------------------------------------------------------- #
@router.get("/evaluasi")
def list_evaluasi(_=Depends(require_eval)):
    return pm.list_evaluasi()


@router.get("/evaluasi/content", response_class=HTMLResponse)
def evaluasi_content(pic: str = Query("", description="Nama PIC"), _=Depends(require_eval)):
    try:
        return HTMLResponse(pm.evaluasi_html(pic))
    except FileNotFoundError as e:
        raise HTTPException(404, str(e))


@router.get("/export")
def export_xlsx(
    section: str = "", status: str = "", pic: str = "",
    bendera: str = "", triwulan: str = "",
    q: str = "", _=Depends(require_pm),
):
    """Ekspor daftar proyek terfilter + ringkasan triwulan ke XLSX."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    try:
        data = pm.list_projects(
            section=section, status=status, pic=pic,
            bendera=bendera, triwulan=triwulan, q=q,
        )
        summ = pm.summary()
    except FileNotFoundError as e:
        raise HTTPException(404, str(e))

    wb = openpyxl.Workbook()
    head_fill = PatternFill("solid", fgColor="1E3A8A")
    head_font = Font(bold=True, color="FFFFFF")

    def style_header(ws, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(1, c)
            cell.fill = head_fill
            cell.font = head_font

    # Sheet 1: daftar proyek.
    ws = wb.active
    ws.title = "Daftar Proyek"
    cols = [
        ("No", "no"), ("Section", "section"), ("Instansi", "instansi"),
        ("Nama Pekerjaan", "nama_pekerjaan"), ("Bendera", "bendera"),
        ("PIC", "pic_names"), ("Status", "status"), ("Progress %", "progress"),
        ("Triwulan", "triwulan"), ("Mulai Kontrak", "mulai_kontrak"),
        ("Akhir Kontrak", "akhir_kontrak"), ("Nilai Project", "nilai_project"),
        ("Nilai Pencapaian", "nilai_pencapaian"), ("Nilai Pencairan", "nilai_pencairan"),
        ("Keterangan", "keterangan"),
    ]
    ws.append([h for h, _k in cols])
    for p in data["projects"]:
        row = []
        for _h, k in cols:
            v = p.get(k)
            if k == "pic_names":
                v = ", ".join(v or [])
            elif k == "progress":
                v = round((v or 0) * 100, 1)
            row.append(v)
        ws.append(row)
    style_header(ws, len(cols))

    # Sheet 2: ringkasan triwulan + evaluasi.
    ws2 = wb.create_sheet("Ringkasan Triwulan")
    tcols = ["Triwulan", "Jumlah", "Done", "Progress", "Nilai Project",
             "Nilai Pencapaian", "Pencapaian %", "Avg Progress %", "Overdue", "Evaluasi"]
    ws2.append(tcols)
    for t in summ["triwulan"]:
        ws2.append([
            t["triwulan"], t["count"], t["done"], t["progress"],
            t["nilai_project"], t["nilai_pencapaian"], t["achievement_rate"],
            t["avg_progress"], t["overdue"], t["evaluasi"],
        ])
    style_header(ws2, len(tcols))
    ws2.column_dimensions["J"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = "project-management-summary.xlsx"
    return StreamingResponse(
        buf, media_type=_XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
