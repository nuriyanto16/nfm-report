"""Modul Project Management: parsing & agregasi "Project List MST 2026".

Sumber data: file XLSX `LAPORAN/00. Project List MST 2026.xlsx`, sheet
"Project List 2026". Sheet ini dipakai untuk:
  * monitoring proyek (status, progress, PIC, bendera, keuangan),
  * ringkasan pencapaian & evaluasi per triwulan (TW I–IV) berdasarkan
    tanggal Akhir Kontrak.

Struktur sheet (header di baris 4):
  A No | B Instansi | C Nama Pekerjaan | D Bendera | E Paket | F Lama |
  G Mulai Kontrak | H Akhir Kontrak | I Nilai Project | J DPP | K Pph23 |
  L DPP Nilai Lain | M Ppn 12% | N Pph KI | O Estimasi Pencairan | P PIC |
  Q Porsi | R Nilai Pengerjaan 2026 | S Nilai Pengerjaan 2025 |
  T Progress 2026 | U Progress 2025 | V Total Progres | W Nilai Pencapaian 2026 |
  X Nilai Pencairan 2026 | Y Nilai Pencairan 2025 | Z Keterangan
  AC (29) status teks: PROGRESS / DONE

Baris section (mis. "PROJECT BERJALAN 2026") = kolom A berisi teks & kolom B
kosong. Sebuah proyek bisa punya beberapa baris (multi-PIC): baris lanjutan
punya No/Instansi kosong tetapi kolom PIC terisi. Baris "Total" (subtotal)
dilewati.
"""
from __future__ import annotations

import os
from datetime import date, datetime
from pathlib import Path
from threading import RLock
from typing import Dict, List, Optional

import openpyxl

SHEET_NAME = "Project List 2026"
HEADER_ROW = 4
_DATA_START = HEADER_ROW + 1

# Kolom (1-based) yang dipakai.
C_NO, C_INSTANSI, C_NAMA, C_BENDERA, C_PAKET, C_LAMA = 1, 2, 3, 4, 5, 6
C_MULAI, C_AKHIR, C_NILAI = 7, 8, 9
C_ESTIMASI = 15
C_PIC, C_PORSI = 16, 17
C_PROG_2026, C_TOTAL_PROG = 20, 22
C_PENCAPAIAN, C_PENCAIRAN_2026 = 23, 24
C_KETERANGAN = 26
C_STATUS = 29

_ROMAN = {1: "I", 2: "II", 3: "III", 4: "IV"}
_lock = RLock()
_cache: Dict[str, object] = {"mtime": None, "data": None}


# --------------------------------------------------------------------------- #
# Lokasi file
# --------------------------------------------------------------------------- #
def _source_path() -> Path:
    env = os.environ.get("PROJECT_LIST_PATH")
    if env:
        return Path(env)
    # <repo>/backend/app/core/project_mgmt.py -> parents[3] = <repo>
    repo = Path(__file__).resolve().parents[3]
    return repo / "LAPORAN" / "00. Project List MST 2026.xlsx"


# --------------------------------------------------------------------------- #
# Util nilai sel
# --------------------------------------------------------------------------- #
def _num(v) -> Optional[float]:
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def _str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _as_date(v) -> Optional[date]:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _quarter_of(d: Optional[date]) -> Optional[int]:
    if not d:
        return None
    return (d.month - 1) // 3 + 1


def _tw_key(d: Optional[date]) -> str:
    q = _quarter_of(d)
    if not q or not d:
        return "Tanpa Tanggal"
    return f"TW {_ROMAN[q]} {d.year}"


# --------------------------------------------------------------------------- #
# Parsing
# --------------------------------------------------------------------------- #
def _is_section(a, b) -> bool:
    return isinstance(a, str) and a.strip() != "" and b is None


def _norm_status(raw: str, total_prog: Optional[float], section: str) -> str:
    s = (raw or "").strip().upper()
    if s in ("DONE", "SELESAI"):
        return "Done"
    if s in ("PROGRESS", "PROGRES", "ON PROGRESS"):
        return "Progress"
    # Turunan bila kolom status kosong.
    if total_prog is not None:
        return "Done" if total_prog >= 0.999 else "Progress"
    if "UPCOMING" in section.upper():
        return "Upcoming"
    if "MAINTENANCE" in section.upper():
        return "Maintenance"
    return "Tanpa Status"


def _parse() -> Dict:
    path = _source_path()
    if not path.exists():
        raise FileNotFoundError(f"File sumber tidak ditemukan: {path}")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{SHEET_NAME}' tidak ada di {path.name}")
    ws = wb[SHEET_NAME]

    projects: List[Dict] = []
    section = ""
    cur: Optional[Dict] = None

    def cell(row, col):
        return row[col - 1] if col - 1 < len(row) else None

    for r_idx, row in enumerate(ws.iter_rows(min_row=_DATA_START, values_only=True), start=_DATA_START):
        a = cell(row, C_NO)
        b = cell(row, C_INSTANSI)

        # Baris section.
        if _is_section(a, b):
            section = a.strip()
            cur = None
            continue

        no = _num(a)
        pic = _str(cell(row, C_PIC))
        # Baris "Total"/subtotal: tidak ada No, tidak ada PIC -> lewati.
        if no is None and not pic:
            cur = None
            continue

        # Baris utama proyek: ada No & Instansi.
        if no is not None and b is not None:
            cur = {
                "no": no,
                "section": section,
                "instansi": _str(b),
                "nama_pekerjaan": _str(cell(row, C_NAMA)),
                "bendera": _str(cell(row, C_BENDERA)) or "-",
                "paket": _str(cell(row, C_PAKET)),
                "lama_pekerjaan": _str(cell(row, C_LAMA)),
                "mulai_kontrak": _as_date(cell(row, C_MULAI)),
                "akhir_kontrak": _as_date(cell(row, C_AKHIR)),
                "nilai_project": _num(cell(row, C_NILAI)) or 0.0,
                "estimasi_pencairan": _num(cell(row, C_ESTIMASI)) or 0.0,
                "total_progress": _num(cell(row, C_TOTAL_PROG)),
                "nilai_pencapaian": _num(cell(row, C_PENCAPAIAN)) or 0.0,
                "nilai_pencairan": _num(cell(row, C_PENCAIRAN_2026)) or 0.0,
                "keterangan": _str(cell(row, C_KETERANGAN)),
                "_status_raw": _str(cell(row, C_STATUS)),
                "pics": [],
                "_row": r_idx,
            }
            projects.append(cur)

        # Alokasi PIC (baris utama maupun lanjutan).
        if cur is not None and pic:
            porsi = _num(cell(row, C_PORSI))
            prog = _num(cell(row, C_PROG_2026))
            pencapaian = _num(cell(row, C_PENCAPAIAN))
            if pic not in [p["pic"] for p in cur["pics"]]:
                cur["pics"].append(
                    {
                        "pic": pic,
                        "porsi": porsi,
                        "progress": prog,
                        "nilai_pencapaian": pencapaian or 0.0,
                    }
                )
            # Total pencapaian proyek = akumulasi lintas baris PIC.
            if r_idx != cur["_row"]:
                cur["nilai_pencapaian"] += pencapaian or 0.0

    wb.close()

    # Finalisasi tiap proyek.
    for p in projects:
        p["status"] = _norm_status(p.pop("_status_raw"), p["total_progress"], p["section"])
        # Progress efektif (0..1). Bila kolom total kosong, rata-rata dari PIC.
        tp = p["total_progress"]
        if tp is None and p["pics"]:
            vals = [x["progress"] for x in p["pics"] if x["progress"] is not None]
            tp = sum(vals) / len(vals) if vals else None
        p["progress"] = round((tp or 0.0), 4)
        p["pic_names"] = [x["pic"] for x in p["pics"]]
        p["triwulan"] = _tw_key(p["akhir_kontrak"])
        p["mulai_kontrak"] = p["mulai_kontrak"].isoformat() if p["mulai_kontrak"] else None
        p["akhir_kontrak_date"] = p["akhir_kontrak"]  # simpan utk hitung risiko
        p["akhir_kontrak"] = p["akhir_kontrak"].isoformat() if p["akhir_kontrak"] else None
        p.pop("total_progress", None)
        p.pop("_row", None)

    return {"projects": projects, "source": path.name}


# --------------------------------------------------------------------------- #
# Cache berbasis mtime file (invalidasi otomatis bila file diperbarui)
# --------------------------------------------------------------------------- #
def _load(force: bool = False) -> Dict:
    path = _source_path()
    mtime = path.stat().st_mtime if path.exists() else None
    with _lock:
        if not force and _cache["data"] is not None and _cache["mtime"] == mtime:
            return _cache["data"]  # type: ignore[return-value]
        data = _parse()
        _cache["data"] = data
        _cache["mtime"] = mtime
        return data


def refresh() -> None:
    with _lock:
        _ds_cache.clear()
    _load(force=True)


# --------------------------------------------------------------------------- #
# Dataset tambahan (Load Mapping, Target bulanan, TW per divisi) — fokus 2026.
# Cache per-dataset berbasis mtime file yang sama; di-reset oleh refresh().
# --------------------------------------------------------------------------- #
_ds_cache: Dict[str, Dict] = {}

LOAD_SHEET = "Load Mapping Tim 2026"
TARGET_SHEETS = [
    ("Januari", "TARGET JAN 2026"),
    ("Februari", "TARGET FEB 2026"),
    ("Maret", "TARGET MAR 2026"),
    ("April", "TARGET APR 2026"),
    ("Mei", "TARGET MEI 2026"),
    ("Juni", "TARGET JUNI 2026"),
]
TW_SHEETS = [
    ("PRODUKSI", "TW II 2026 - PRODUKSI"),
    ("BIZDEV", "TW II 2026 - BIZDEV"),
    ("HR-GA-FIN", "TW II 2026 - HR GA FIN"),
]


def _cached(key: str, builder) -> Dict:
    path = _source_path()
    mtime = path.stat().st_mtime if path.exists() else None
    with _lock:
        e = _ds_cache.get(key)
        if e and e["mtime"] == mtime:
            return e["data"]
        data = builder()
        _ds_cache[key] = {"mtime": mtime, "data": data}
        return data


def _open():
    path = _source_path()
    if not path.exists():
        raise FileNotFoundError(f"File sumber tidak ditemukan: {path}")
    return openpyxl.load_workbook(path, data_only=True)


def _norm_hdr(v) -> str:
    return " ".join(str(v).split()).upper() if v is not None else ""


def _header_row(ws, scan: int = 12) -> Optional[int]:
    for r in range(1, scan + 1):
        if _norm_hdr(ws.cell(r, 1).value).rstrip(".") == "NO":
            return r
    return None


def _colmap(ws, hr: int, spec: Dict[str, list]) -> Dict[str, int]:
    """Petakan nama-logis -> nomor kolom dgn mencocokkan teks header (predikat)."""
    out: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        u = _norm_hdr(ws.cell(hr, c).value)
        if not u:
            continue
        for key, preds in spec.items():
            if key in out:
                continue
            if any(p(u) for p in preds):
                out[key] = c
    return out


def _norm_status_text(s: str) -> str:
    u = (s or "").strip().lower()
    if not u:
        return "Tanpa Status"
    if u in ("done", "selesai", "achieve", "achieved", "tercapai"):
        return "Done"
    if "not achieve" in u or "not achieved" in u:
        return "Not Achieve"
    if "progress" in u or "progres" in u:
        return "In Progress"
    if "adjust" in u:
        return "Adjustment"
    if "to do" in u or "todo" in u:
        return "To Do"
    return s.strip().title()


# ------------------------------- Load Mapping ------------------------------- #
def _build_load_mapping() -> Dict:
    wb = _open()
    if LOAD_SHEET not in wb.sheetnames:
        return {"members": [], "projects": [], "member_totals": [], "counts": {"total": 0, "done": 0}}
    ws = wb[LOAD_SHEET]
    hr = _header_row(ws) or 3
    spec = {
        "instansi": [lambda u: u == "INSTANSI"],
        "nama": [lambda u: "NAMA" in u or "PEKERJAAN" in u and "PROGRESS" not in u],
        "progress": [lambda u: u.startswith("PROGRESS")],
        "tahun": [lambda u: u == "TAHUN"],
        "pic": [lambda u: "PIC" in u],
        "total": [lambda u: "TOTAL" in u and "MAPPING" in u],
    }
    m = _colmap(ws, hr, spec)
    start = (m.get("total") or m.get("pic") or 6) + 1
    members = []
    for c in range(start, ws.max_column + 1):
        name = _str(ws.cell(hr, c).value)
        if name:
            members.append({"name": name, "col": c})

    projects = []
    for r in range(hr + 1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, m.get("instansi", 2)).value
        if _num(a) is None:
            continue
        if b in (None, ""):
            continue
        alloc = {}
        for mem in members:
            v = _num(ws.cell(r, mem["col"]).value)
            if v:
                alloc[mem["name"]] = round(v, 4)
        projects.append({
            "no": _num(a),
            "instansi": _str(b),
            "nama": _str(ws.cell(r, m["nama"]).value) if "nama" in m else "",
            "pic_utama": _str(ws.cell(r, m["pic"]).value) if "pic" in m else "",
            "tahun": _str(ws.cell(r, m["tahun"]).value) if "tahun" in m else "",
            "progress": round(_num(ws.cell(r, m["progress"]).value) or 0.0, 4) if "progress" in m else 0.0,
            "total_mapping": round(_num(ws.cell(r, m["total"]).value) or 0.0, 4) if "total" in m else 0.0,
            "alloc": alloc,
        })

    totals: Dict[str, Dict] = {}
    for p in projects:
        for name, v in p["alloc"].items():
            t = totals.setdefault(name, {"member": name, "total_load": 0.0, "projects": 0})
            t["total_load"] += v
            t["projects"] += 1
    member_totals = sorted(
        ({**t, "total_load": round(t["total_load"], 3)} for t in totals.values()),
        key=lambda x: -x["total_load"],
    )
    done = sum(1 for p in projects if p["progress"] >= 0.999)
    return {
        "members": [m2["name"] for m2 in members],
        "projects": projects,
        "member_totals": member_totals,
        "counts": {"total": len(projects), "done": done},
    }


def load_mapping() -> Dict:
    return _cached("load_mapping", _build_load_mapping)


# ------------------------------- Target bulanan ----------------------------- #
def _build_targets() -> Dict:
    wb = _open()
    months = []
    overall: Dict[str, int] = {}
    for label, sheet in TARGET_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hr = _header_row(ws) or 4
        m = _colmap(ws, hr, {
            "instansi": [lambda u: u == "INSTANSI"],
            "kegiatan": [lambda u: "PROJECT" in u or "KEGIATAN" in u],
            "target": [lambda u: u == "TARGET" or u.startswith("TARGET ")],
            "status": [lambda u: "STATUS" in u],
            "keterangan": [lambda u: u.startswith("KETERANGAN")],
        })
        items = []
        by_status: Dict[str, int] = {}
        group = ""
        for r in range(hr + 1, ws.max_row + 1):
            a = ws.cell(r, 1).value
            b = ws.cell(r, m.get("instansi", 2)).value
            if _num(a) is None:
                if isinstance(a, str) and a.strip() and (b in (None, "")):
                    group = a.strip()
                continue
            status = _norm_status_text(_str(ws.cell(r, m["status"]).value)) if "status" in m else "Tanpa Status"
            items.append({
                "group": group,
                "instansi": _str(b),
                "kegiatan": _str(ws.cell(r, m["kegiatan"]).value) if "kegiatan" in m else "",
                "target": _str(ws.cell(r, m["target"]).value) if "target" in m else "",
                "status": status,
                "keterangan": _str(ws.cell(r, m["keterangan"]).value) if "keterangan" in m else "",
            })
            by_status[status] = by_status.get(status, 0) + 1
            overall[status] = overall.get(status, 0) + 1
        months.append({"month": label, "sheet": sheet, "total": len(items),
                       "by_status": by_status, "items": items})
    return {"months": months, "by_status": overall}


def targets() -> Dict:
    return _cached("targets", _build_targets)


# ------------------------------- TW per divisi ------------------------------ #
def _build_tw() -> Dict:
    wb = _open()
    divisions = []
    for label, sheet in TW_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hr = _header_row(ws)
        if hr is None:
            continue
        m = _colmap(ws, hr, {
            "instansi": [lambda u: u == "INSTANSI"],
            "kegiatan": [lambda u: "KEGIATAN" in u or "PROJECT" in u],
            "tahun": [lambda u: u == "TAHUN"],
            "pic": [lambda u: u == "PIC"],
            "target_tw": [lambda u: u.startswith("TARGET TRIWULAN")],
            "progress": [lambda u: u.startswith("PROGRESS")],
            "status": [lambda u: u.startswith("STATUS")],
            "keterangan": [lambda u: u.startswith("KETERANGAN")],
        })
        items = []
        by_status: Dict[str, int] = {}
        group = ""
        for r in range(hr + 1, ws.max_row + 1):
            a = ws.cell(r, 1).value
            if _num(a) is None:
                if isinstance(a, str) and a.strip():
                    group = a.strip()
                continue
            prog = _num(ws.cell(r, m["progress"]).value) if "progress" in m else None
            if prog is not None:
                status = "Done" if prog >= 0.999 else "Progress"
            elif "status" in m:
                status = _norm_status_text(_str(ws.cell(r, m["status"]).value))
            else:
                status = "Tanpa Status"
            items.append({
                "group": group,
                "no": _num(a),
                "instansi": _str(ws.cell(r, m["instansi"]).value) if "instansi" in m else "",
                "kegiatan": _str(ws.cell(r, m["kegiatan"]).value) if "kegiatan" in m else "",
                "pic": _str(ws.cell(r, m["pic"]).value) if "pic" in m else "",
                "tahun": _str(ws.cell(r, m["tahun"]).value) if "tahun" in m else "",
                "target_tw": _str(ws.cell(r, m["target_tw"]).value) if "target_tw" in m else "",
                "progress": round(prog, 4) if prog is not None else None,
                "status": status,
                "keterangan": _str(ws.cell(r, m["keterangan"]).value) if "keterangan" in m else "",
            })
            by_status[status] = by_status.get(status, 0) + 1
        done = by_status.get("Done", 0)
        divisions.append({"division": label, "sheet": sheet, "total": len(items),
                          "done": done, "by_status": by_status, "items": items})
    return {"divisions": divisions}


def tw() -> Dict:
    return _cached("tw", _build_tw)


# --------------------------------------------------------------------------- #
# Query publik
# --------------------------------------------------------------------------- #
def _public_project(p: Dict) -> Dict:
    return {k: v for k, v in p.items() if not k.startswith("_") and k != "akhir_kontrak_date"}


def filters() -> Dict[str, List[str]]:
    data = _load()
    sections, statuses, pics, benderas, tws = set(), set(), set(), set(), set()
    for p in data["projects"]:
        sections.add(p["section"])
        statuses.add(p["status"])
        benderas.add(p["bendera"])
        tws.add(p["triwulan"])
        for name in p["pic_names"]:
            pics.add(name)
    return {
        "sections": sorted(sections),
        "statuses": sorted(statuses),
        "pics": sorted(pics),
        "benderas": sorted(benderas),
        "triwulan": sorted(tws, key=lambda t: (t == "Tanpa Tanggal", t)),
    }


def _match(p: Dict, *, section, status, pic, bendera, triwulan, q) -> bool:
    if section and p["section"] != section:
        return False
    if status and p["status"] != status:
        return False
    if bendera and p["bendera"] != bendera:
        return False
    if triwulan and p["triwulan"] != triwulan:
        return False
    if pic and pic not in p["pic_names"]:
        return False
    if q:
        hay = f"{p['instansi']} {p['nama_pekerjaan']}".lower()
        if q.lower() not in hay:
            return False
    return True


def list_projects(
    *, section: str = "", status: str = "", pic: str = "",
    bendera: str = "", triwulan: str = "", q: str = "",
) -> Dict:
    data = _load()
    rows = [
        _public_project(p)
        for p in data["projects"]
        if _match(p, section=section, status=status, pic=pic,
                  bendera=bendera, triwulan=triwulan, q=q)
    ]
    return {"projects": rows, "count": len(rows), "source": data["source"]}


def _pct(part: float, whole: float) -> float:
    return round((part / whole) * 100, 1) if whole else 0.0


def _evaluasi_tw(tw: str, s: Dict, today: date) -> str:
    """Rangkuman evaluasi naratif per triwulan (rule-based, Bahasa Indonesia)."""
    parts: List[str] = []
    total = s["count"]
    if total == 0:
        return "Belum ada proyek yang jatuh tempo pada triwulan ini."
    done_pct = _pct(s["done"], total)
    parts.append(
        f"{total} proyek berakhir pada {tw}; {s['done']} selesai ({done_pct}%), "
        f"{s['progress']} berjalan."
    )
    ach = s["achievement_rate"]
    if ach >= 90:
        parts.append(f"Pencapaian nilai sangat baik ({ach}% dari target).")
    elif ach >= 70:
        parts.append(f"Pencapaian nilai cukup ({ach}% dari target), masih ada ruang perbaikan.")
    else:
        parts.append(f"Pencapaian nilai rendah ({ach}% dari target) — perlu perhatian.")
    if s["overdue"] > 0:
        parts.append(
            f"⚠️ {s['overdue']} proyek melewati akhir kontrak namun progres belum 100% "
            "— prioritaskan penyelesaian & penagihan."
        )
    else:
        parts.append("Tidak ada proyek yang overdue.")
    return " ".join(parts)


def summary(today: Optional[date] = None) -> Dict:
    data = _load()
    today = today or date.today()
    projects = data["projects"]

    by_status: Dict[str, int] = {}
    by_section: Dict[str, int] = {}
    by_bendera: Dict[str, int] = {}
    by_pic: Dict[str, Dict] = {}
    tw_map: Dict[str, Dict] = {}

    tot_nilai = tot_estimasi = tot_pencapaian = tot_pencairan = 0.0
    prog_sum = prog_n = 0
    overdue: List[Dict] = []
    due_soon: List[Dict] = []

    for p in projects:
        by_status[p["status"]] = by_status.get(p["status"], 0) + 1
        by_section[p["section"]] = by_section.get(p["section"], 0) + 1
        by_bendera[p["bendera"]] = by_bendera.get(p["bendera"], 0) + 1
        for name in p["pic_names"]:
            b = by_pic.setdefault(name, {"total": 0, "done": 0, "nilai_pencapaian": 0.0})
            b["total"] += 1
            if p["status"] == "Done":
                b["done"] += 1
        by_pic_pencapaian_add(by_pic, p)

        tot_nilai += p["nilai_project"]
        tot_estimasi += p["estimasi_pencairan"]
        tot_pencapaian += p["nilai_pencapaian"]
        tot_pencairan += p["nilai_pencairan"]
        if p["nilai_project"] > 0 or p["progress"] > 0:
            prog_sum += p["progress"]
            prog_n += 1

        # Triwulan (berdasarkan Akhir Kontrak).
        tw = p["triwulan"]
        t = tw_map.setdefault(tw, {
            "count": 0, "done": 0, "progress": 0,
            "nilai_project": 0.0, "nilai_pencapaian": 0.0, "nilai_pencairan": 0.0,
            "progress_sum": 0.0, "progress_n": 0, "overdue": 0,
        })
        t["count"] += 1
        if p["status"] == "Done":
            t["done"] += 1
        elif p["status"] == "Progress":
            t["progress"] += 1
        t["nilai_project"] += p["nilai_project"]
        t["nilai_pencapaian"] += p["nilai_pencapaian"]
        t["nilai_pencairan"] += p["nilai_pencairan"]
        if p["nilai_project"] > 0 or p["progress"] > 0:
            t["progress_sum"] += p["progress"]
            t["progress_n"] += 1

        # Risiko: akhir kontrak lewat tapi progres < 100%.
        ad = p.get("akhir_kontrak_date")
        if ad and p["progress"] < 0.999:
            if ad < today:
                t["overdue"] += 1
                overdue.append(_risk_row(p, (today - ad).days))
            elif (ad - today).days <= 30:
                due_soon.append(_risk_row(p, -(ad - today).days))

    # Finalisasi triwulan + evaluasi.
    triwulan = []
    for tw in sorted(tw_map, key=lambda k: (k == "Tanpa Tanggal", k)):
        t = tw_map[tw]
        t["avg_progress"] = round((t.pop("progress_sum") / t["progress_n"]) * 100, 1) if t.get("progress_n") else 0.0
        t.pop("progress_n", None)
        t["achievement_rate"] = _pct(t["nilai_pencapaian"], t["nilai_project"])
        t["evaluasi"] = _evaluasi_tw(tw, t, today)
        triwulan.append({"triwulan": tw, **t})

    pic_rows = sorted(
        (
            {"pic": k, **v, "done_rate": _pct(v["done"], v["total"])}
            for k, v in by_pic.items()
        ),
        key=lambda x: -x["total"],
    )

    return {
        "source": data["source"],
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "totals": {
            "projects": len(projects),
            "nilai_project": tot_nilai,
            "estimasi_pencairan": tot_estimasi,
            "nilai_pencapaian": tot_pencapaian,
            "nilai_pencairan": tot_pencairan,
            "avg_progress": round((prog_sum / prog_n) * 100, 1) if prog_n else 0.0,
            "achievement_rate": _pct(tot_pencapaian, tot_nilai),
            "pencairan_rate": _pct(tot_pencairan, tot_nilai),
            "done": by_status.get("Done", 0),
            "overdue": len(overdue),
        },
        "by_status": by_status,
        "by_section": by_section,
        "by_bendera": dict(sorted(by_bendera.items(), key=lambda x: -x[1])),
        "by_pic": pic_rows,
        "triwulan": triwulan,
        "alerts": {
            "overdue": sorted(overdue, key=lambda x: -x["days"]),
            "due_soon": sorted(due_soon, key=lambda x: x["days"]),
        },
    }


def by_pic_pencapaian_add(by_pic: Dict[str, Dict], p: Dict) -> None:
    """Bagikan nilai pencapaian proyek ke tiap PIC-nya (dibagi rata)."""
    names = p["pic_names"]
    if not names:
        return
    share = p["nilai_pencapaian"] / len(names)
    for name in names:
        by_pic[name]["nilai_pencapaian"] += share


# --------------------------------------------------------------------------- #
# Evaluasi Triwulan (dokumen HTML per PIC yang di-embed)
# --------------------------------------------------------------------------- #
# File berpola "<judul>_Tim_<PIC>.html" di folder LAPORAN. Nama PIC diambil dari
# bagian setelah "_Tim_". Untuk sementara baru ada satu (Tim Nuriyanto); struktur
# ini memudahkan penambahan file per-PIC berikutnya tanpa ubah kode.
_EVALUASI_GLOB = "*_Tim_*.html"


def _evaluasi_dir() -> Path:
    env = os.environ.get("EVALUASI_DIR")
    if env:
        return Path(env)
    return _source_path().parent


def _title_of(path: Path) -> str:
    try:
        head = path.read_text(encoding="utf-8", errors="ignore")[:4000]
    except OSError:
        return path.stem
    import re

    m = re.search(r"<title>(.*?)</title>", head, re.IGNORECASE | re.DOTALL)
    return m.group(1).strip() if m else path.stem


def _pic_of(path: Path) -> str:
    stem = path.stem
    part = stem.split("_Tim_")[-1] if "_Tim_" in stem else stem
    return part.replace("_", " ").strip()


def list_evaluasi() -> Dict:
    d = _evaluasi_dir()
    items = []
    if d.exists():
        for f in sorted(d.glob(_EVALUASI_GLOB)):
            items.append({"pic": _pic_of(f), "file": f.name, "title": _title_of(f)})
    return {"items": items}


def evaluasi_html(pic: str) -> str:
    d = _evaluasi_dir()
    if not d.exists():
        raise FileNotFoundError("Folder dokumen evaluasi tidak ditemukan")
    want = (pic or "").strip().lower()
    match: Optional[Path] = None
    for f in sorted(d.glob(_EVALUASI_GLOB)):
        if not want or _pic_of(f).lower() == want:
            match = f
            break
    if match is None:
        raise FileNotFoundError(f"Dokumen evaluasi untuk PIC '{pic}' tidak ditemukan")
    return match.read_text(encoding="utf-8", errors="ignore")


def _risk_row(p: Dict, days: int) -> Dict:
    return {
        "no": p["no"],
        "instansi": p["instansi"],
        "nama_pekerjaan": p["nama_pekerjaan"],
        "pic": ", ".join(p["pic_names"]) or "-",
        "bendera": p["bendera"],
        "akhir_kontrak": p["akhir_kontrak"],
        "progress": p["progress"],
        "status": p["status"],
        "days": days,  # + = hari lewat jatuh tempo; - = sisa hari menuju jatuh tempo
        "triwulan": p["triwulan"],
    }
