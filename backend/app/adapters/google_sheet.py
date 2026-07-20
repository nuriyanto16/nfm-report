"""Adapter sumber Google Sheet (publik, di-share tanpa auth).

Men-download seluruh workbook lewat endpoint export xlsx, lalu mem-parsing
setiap tab bulanan sesuai layout di sources.yaml menjadi list[TaskRecord].
"""
from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Any, Dict, List, Optional

import httpx
import openpyxl

from app.adapters.base import SourceAdapter
from app.core.models import TANPA_STATUS, TaskRecord, canonical_status

EXPORT_URL = "https://docs.google.com/spreadsheets/d/{id}/export?format=xlsx"


def _clean(v: Any) -> Optional[str]:
    """Normalisasi nilai sel string: trim, buang spasi ganda, '' -> None."""
    if v is None:
        return None
    if isinstance(v, str):
        s = " ".join(v.split()).strip()
        return s or None
    return v


def _text(v: Any) -> Optional[str]:
    """Seperti _clean tapi nilai numerik dipaksa jadi string (utk field teks)."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = " ".join(str(v).split()).strip()
    return s or None


def _as_dt(v: Any) -> Optional[datetime]:
    if isinstance(v, datetime):
        return v
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
    return None


def _as_float(v: Any) -> Optional[float]:
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _norm_header(s: Any) -> str:
    """Normalisasi teks header untuk pencocokan: upper, buang /,\\n, spasi ganda."""
    if s is None:
        return ""
    t = str(s).replace("/", " ").replace("\n", " ").replace("\\", " ")
    return " ".join(t.split()).strip().upper()


# Alias header -> field TaskRecord (untuk mode deteksi kolom otomatis).
# Urutan alias dari yang paling spesifik agar pencocokan tidak salah kolom.
DEFAULT_HEADER_ALIASES: Dict[str, List[str]] = {
    "no": ["NO"],
    "nomor_ticket": ["NOMOR TICKET", "NO TICKET"],
    "issue": ["ISSUE", "URAIAN", "KEGIATAN"],
    "path_menu": ["PATH MENU APLIKASI", "PATH MENU"],
    "tgl_request": ["TANGGAL REQUEST ISSUE", "TANGGAL REQUEST", "TGL REQUEST"],
    "tgl_mulai": ["TANGGAL MULAI PEKERJAAN", "TANGGAL MULAI PENGERJAAN",
                  "TANGGAL MULAI", "TGL MULAI"],
    "tgl_estimasi": ["TANGGAL ESTIMASI SELESAI", "TANGGAL ESTIMASI", "TGL ESTIMASI"],
    "tgl_selesai": ["TANGGAL SELESAI", "TGL SELESAI"],
    "request_by": ["REQUEST BY", "REQUESTED BY"],
    "pic": ["PIC"],
    "kategori": ["KATEGORI"],
    "status": ["STATUS"],
    "priority": ["PRIORITY", "PRIORITAS"],
    "status_deploy": ["STATUS DEPLOY"],
    "last_update": ["LAST UPDATE CHECK", "LAST UPDATE"],
    "keterangan": ["KETERANGAN", "CATATAN"],
}

_SECTION_RE = re.compile(r"^\s*\d+\.\s+\S")


class GoogleSheetAdapter(SourceAdapter):
    def __init__(self, source: Dict[str, Any]):
        super().__init__(source)
        self.spreadsheet_id = source.get("spreadsheet_id")
        self.local_path = source.get("local_path")  # untuk test/offline
        self.sheet_pattern = re.compile(source.get("sheet_pattern", r".*"))
        self.layout = source.get("layout", {})

    # --- pengambilan data mentah -------------------------------------------
    def _load_bytes(self) -> bytes:
        if self.local_path:
            with open(self.local_path, "rb") as f:
                return f.read()
        url = EXPORT_URL.format(id=self.spreadsheet_id)
        resp = httpx.get(url, follow_redirects=True, timeout=60.0)
        resp.raise_for_status()
        return resp.content

    def fetch(self) -> List[TaskRecord]:
        wb = openpyxl.load_workbook(
            io.BytesIO(self._load_bytes()), data_only=True, read_only=True
        )
        records: List[TaskRecord] = []
        for name in wb.sheetnames:
            if self.sheet_pattern.match(name):
                records.extend(self._parse_sheet(wb[name], name))
        return records

    # --- parsing per tab bulanan -------------------------------------------
    def _parse_sheet(self, ws, sheet_name: str) -> List[TaskRecord]:
        # Mode deteksi kolom otomatis (per-tab, dari baris header) — dipakai
        # bila layout punya detect_columns: true. Mendukung tab dgn layout
        # kolom berbeda dalam satu spreadsheet.
        if self.layout.get("detect_columns"):
            return self._parse_sheet_detect(ws, sheet_name)
        col = self.layout["columns"]
        data_start = self.layout.get("data_start_row", 4)
        section_markers: Dict[str, str] = self.layout.get("section_markers", {})
        stop_markers = {s.strip().lower()
                        for s in self.layout.get("stop_markers", [])}
        max_col = max(col.values()) + 1

        out: List[TaskRecord] = []
        current_aplikasi = ""
        seen_data = False  # untuk membedakan judul atas vs blok legacy bawah

        rows = ws.iter_rows(min_row=1, max_col=max_col, values_only=True)
        for idx, row in enumerate(rows, start=1):
            a = row[col["no"]] if col["no"] < len(row) else None

            # Baris penanda (col A berupa teks non-numerik).
            if isinstance(a, str):
                marker = " ".join(a.split()).strip()
                low = marker.lower()
                # Stop saat blok legacy non-SIMS dimulai (judul muncul lagi
                # SETELAH data SIMPEL terbaca).
                if seen_data and low in stop_markers:
                    break
                if marker in section_markers:
                    current_aplikasi = section_markers[marker]
                continue

            if idx < data_start:
                continue

            no = _as_float(a)
            issue = _clean(self._cell(row, col, "issue"))
            # Baris data valid: NO numerik DAN ISSUE berupa teks tidak kosong.
            # ISSUE numerik = tabel mini-recap (noise) -> di-skip.
            if no is None or not isinstance(issue, str) or not issue:
                continue

            seen_data = True
            raw_status = _clean(self._cell(row, col, "status"))
            out.append(TaskRecord(
                no=no,
                issue=issue,
                path_menu=_clean(self._cell(row, col, "path_menu")),
                tgl_request=_as_dt(self._cell(row, col, "tgl_request")),
                tgl_mulai=_as_dt(self._cell(row, col, "tgl_mulai")),
                tgl_estimasi=_as_dt(self._cell(row, col, "tgl_estimasi")),
                tgl_selesai=_as_dt(self._cell(row, col, "tgl_selesai")),
                request_by=_clean(self._cell(row, col, "request_by")),
                pic=_clean(self._cell(row, col, "pic")),
                kategori=_clean(self._cell(row, col, "kategori")),
                status=canonical_status(raw_status),
                status_raw=raw_status,
                priority=_clean(self._cell(row, col, "priority")),
                status_deploy=_clean(self._cell(row, col, "status_deploy")),
                last_update=_as_dt(self._cell(row, col, "last_update")),
                keterangan=_clean(self._cell(row, col, "keterangan")),
                aplikasi=current_aplikasi,
                source_id=self.source_id,
                source_month=sheet_name,
            ))
        return out

    @staticmethod
    def _cell(row, col: Dict[str, int], key: str):
        i = col.get(key)
        if i is None or i >= len(row):
            return None
        return row[i]

    # --- mode deteksi kolom otomatis ---------------------------------------
    def _detect_columns(self, header_row) -> Dict[str, int]:
        """Bangun peta field->indeks kolom dari baris header (per-tab)."""
        aliases = self.layout.get("field_aliases") or DEFAULT_HEADER_ALIASES
        norm = [_norm_header(h) for h in header_row]
        colmap: Dict[str, int] = {}
        used = set()

        # Pass 1: cocok PERSIS (paling aman).
        for field, al_list in aliases.items():
            for al in al_list:
                na = _norm_header(al)
                for i, h in enumerate(norm):
                    if i in used or not h:
                        continue
                    if h == na:
                        colmap[field] = i
                        used.add(i)
                        break
                if field in colmap:
                    break
        # Pass 2: cocok awalan untuk field yang belum ketemu.
        for field, al_list in aliases.items():
            if field in colmap:
                continue
            for al in al_list:
                na = _norm_header(al)
                for i, h in enumerate(norm):
                    if i in used or not h:
                        continue
                    if h.startswith(na):
                        colmap[field] = i
                        used.add(i)
                        break
                if field in colmap:
                    break
        return colmap

    def _parse_sheet_detect(self, ws, sheet_name: str) -> List[TaskRecord]:
        header_row_no = self.layout.get("header_row", 3)
        section_markers: Dict[str, str] = self.layout.get("section_markers", {})
        stop_markers = {s.strip().lower()
                        for s in self.layout.get("stop_markers", [])}

        # Whitelist status untuk membuang baris tabel rekap mini (status berupa
        # angka / label prioritas seperti Low/Middle/High/Total/#DIV/0!).
        valid_status = {canonical_status(s)
                        for s in self.layout.get("valid_status", [])}

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < header_row_no:
            return []
        colmap = self._detect_columns(rows[header_row_no - 1])
        if "issue" not in colmap:  # tab tanpa kolom ISSUE bukan log task
            return []

        def cell(row, key):
            i = colmap.get(key)
            if i is None or i >= len(row):
                return None
            return row[i]

        out: List[TaskRecord] = []
        current_aplikasi = ""
        seen_data = False

        for idx, row in enumerate(rows, start=1):
            a = row[0] if row else None
            # Baris penanda (col A teks): section / banner / header.
            if isinstance(a, str) and a.strip():
                marker = " ".join(a.split()).strip()
                if seen_data and marker.lower() in stop_markers:
                    break
                if _SECTION_RE.match(marker):
                    current_aplikasi = section_markers.get(marker) or \
                        re.sub(r"^\s*\d+\.\s*", "", marker)
                continue
            if idx <= header_row_no:
                continue

            issue = _clean(cell(row, "issue"))
            if not isinstance(issue, str) or not issue:
                continue

            raw_status = _text(cell(row, "status"))
            status = canonical_status(raw_status)
            # Buang baris rekap: status tak dikenal & bukan kosong.
            if valid_status and status not in valid_status and status != TANPA_STATUS:
                continue

            seen_data = True
            out.append(TaskRecord(
                no=_as_float(cell(row, "no")),
                issue=issue,
                path_menu=_text(cell(row, "path_menu")),
                tgl_request=_as_dt(cell(row, "tgl_request")),
                tgl_mulai=_as_dt(cell(row, "tgl_mulai")),
                tgl_estimasi=_as_dt(cell(row, "tgl_estimasi")),
                tgl_selesai=_as_dt(cell(row, "tgl_selesai")),
                request_by=_text(cell(row, "request_by")),
                pic=_text(cell(row, "pic")),
                kategori=_text(cell(row, "kategori")),
                status=status,
                status_raw=raw_status,
                priority=_text(cell(row, "priority")),
                status_deploy=_text(cell(row, "status_deploy")),
                last_update=_as_dt(cell(row, "last_update")),
                keterangan=_text(cell(row, "keterangan")),
                aplikasi=current_aplikasi,
                source_id=self.source_id,
                source_month=sheet_name,
            ))
        return out
