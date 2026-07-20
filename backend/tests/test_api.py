"""Integration test endpoint + export, memakai sumber lokal sintetis."""
from __future__ import annotations

import io

import openpyxl
import pytest
import yaml
from fastapi.testclient import TestClient


@pytest.fixture
def client(synthetic_xlsx, tmp_path, monkeypatch):
    cfg = {
        "sources": [{
            "id": "test-src",
            "name": "Test Source",
            "type": "google_sheet",
            "local_path": synthetic_xlsx,
            "sheet_pattern": r"^\d{2}\. .+ \d{4}$",
            "layout": {
                "data_start_row": 4,
                "columns": {
                    "no": 0, "issue": 1, "path_menu": 2, "tgl_request": 3,
                    "tgl_mulai": 4, "tgl_estimasi": 5, "tgl_selesai": 6,
                    "request_by": 7, "pic": 8, "kategori": 9, "status": 10,
                    "priority": 11, "status_deploy": 12, "last_update": 13,
                    "keterangan": 15,
                },
                "section_markers": {
                    "1. Pengujian": "Pengujian (SIMPEL)",
                    "2. Kalibrasi": "Kalibrasi",
                },
                "stop_markers": [
                    "CHANGE REQUEST (CR)/FEEDBACK USER",
                    "BULAN DESEMBER 2025",
                ],
            },
        }]
    }
    cfg_path = tmp_path / "sources.yaml"
    cfg_path.write_text(yaml.safe_dump(cfg), encoding="utf-8")
    monkeypatch.setenv("SOURCES_CONFIG", str(cfg_path))
    monkeypatch.setenv("AUTH_DISABLED", "1")  # test ini fokus data, bukan auth

    from app.core import config, repository
    config.reset_config_cache()
    repository.clear_cache()
    from app.main import app
    return TestClient(app)


def test_sources(client):
    r = client.get("/api/sources")
    assert r.status_code == 200
    assert r.json()["sources"][0]["id"] == "test-src"


def test_filters_endpoint(client):
    r = client.get("/api/filters", params={"source": "test-src"})
    assert r.status_code == 200
    body = r.json()
    assert "Done" in body["values"]["status"]
    assert "tgl_request" in body["date_fields"]


def test_report_monthly(client):
    r = client.get("/api/report", params={
        "source": "test-src", "period": "monthly", "month": "2026-06",
    })
    assert r.status_code == 200
    body = r.json()
    assert body["meta"]["period"]["label"] == "Juni 2026"
    # Hanya task Juni dgn tgl_request terisi (3 dari 5: r6, r7, r9; r8 tgl kosong, kalibrasi r13 ada).
    issues = [row["issue"] for row in body["rows"]]
    assert "Bug notifikasi WA" in issues
    assert all("legacy" not in i.lower() for i in issues)


def test_report_filter_status(client):
    r = client.get("/api/report", params={
        "source": "test-src", "period": "monthly", "month": "2026-06",
        "status": "Done",
    })
    assert r.status_code == 200
    assert all(row["status"] == "Done" for row in r.json()["rows"])


def test_report_daily(client):
    r = client.get("/api/report", params={
        "source": "test-src", "period": "daily", "date": "2026-06-03",
    })
    assert r.status_code == 200
    assert [row["issue"] for row in r.json()["rows"]] == ["Bug notifikasi WA"]


def test_export_xlsx(client):
    r = client.get("/api/export", params={
        "source": "test-src", "period": "monthly", "month": "2026-06",
        "format": "xlsx",
    })
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats")
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert "Rekap" in wb.sheetnames and "Summary" in wb.sheetnames
    assert "attachment" in r.headers["content-disposition"]


def test_export_docx(client):
    r = client.get("/api/export", params={
        "source": "test-src", "period": "monthly", "month": "2026-06",
        "format": "docx",
    })
    assert r.status_code == 200
    assert r.content[:2] == b"PK"  # docx = zip


def test_export_pdf_if_libreoffice(client):
    from app.exporters import pdf as pdfmod
    try:
        pdfmod._find_soffice()
    except RuntimeError:
        pytest.skip("LibreOffice tidak tersedia")
    r = client.get("/api/export", params={
        "source": "test-src", "period": "monthly", "month": "2026-06",
        "format": "pdf",
    })
    assert r.status_code == 200
    assert r.content[:5] == b"%PDF-"


def test_refresh(client):
    r = client.post("/api/refresh", params={"source": "test-src"})
    assert r.status_code == 200
    assert r.json()["status"] == "refreshed"
