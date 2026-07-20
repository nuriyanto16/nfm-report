"""Test mode deteksi kolom otomatis (layout berbeda dalam 1 spreadsheet)."""
from __future__ import annotations

import openpyxl
import pytest

from app.adapters.google_sheet import GoogleSheetAdapter


def _hdr(ws, headers, row=3):
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)


@pytest.fixture
def two_layout_xlsx(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "Summary"  # diabaikan (tak match pola)

    # Tab 2025: NO, ISSUE, TGL REQUEST, TGL SELESAI, REQUEST BY, PIC, KATEGORI, STATUS, KETERANGAN
    a = wb.create_sheet("01. Mei 2025")
    a["A1"] = "CHANGE REQUEST (CR)/FEEDBACK USER"
    _hdr(a, ["NO", "ISSUE", "TANGGAL REQUEST", "TANGGAL SELESAI",
             "REQUEST BY", "PIC", "KATEGORI", "STATUS", "KETERANGAN"])
    a["A4"] = "BULAN MEI 2025"
    a["A5"] = "1. Aplikasi CAT Sertifikasi Operator Radio"
    from datetime import datetime
    a.append([])  # row 6 placeholder offset not needed; set explicitly below
    for c, v in enumerate([1, "Pengecekan akses VM", datetime(2025, 5, 6),
                           datetime(2025, 5, 6), "Pa Didik", "Rio", "Support",
                           "Done", "ok"], start=1):
        a.cell(row=6, column=c, value=v)

    # Tab 2026: NO(kosong), NOMOR TICKET, ISSUE, TGL REQUEST, ... STATUS, PRIORITY
    b = wb.create_sheet("06. Juni 2026")
    b["A1"] = "CHANGE REQUEST (CR)/FEEDBACK USER"
    _hdr(b, ["NO", "NOMOR TICKET", "ISSUE", "TANGGAL REQUEST",
             "TANGGAL MULAI PEKERJAAN", "TANGGAL ESTIMASI SELESAI",
             "TANGGAL SELESAI", "REQUEST BY", "PIC", "KATEGORI", "STATUS",
             "PRIORITY"])
    b["A4"] = "BULAN JUNI 2026"
    b["A5"] = "1. Aplikasi CAT Sertifikasi Operator Radio"
    # NO kosong, ticket di kolom B, issue di C
    for c, v in enumerate([None, "01-062026-0001", "Scan Accunetix",
                           datetime(2026, 6, 10), datetime(2026, 6, 10),
                           datetime(2026, 6, 11), datetime(2026, 6, 11),
                           "Pa Bangsawan", "Rio", "Security", "Progress",
                           "High"], start=1):
        b.cell(row=6, column=c, value=v)
    # baris rekap noise: status berupa label prioritas -> harus dibuang
    b["A8"] = "14. Other"
    for c, v in enumerate([None, None, "CAT", None, None, None, None, None,
                           0, None, "Low", None], start=1):
        b.cell(row=9, column=c, value=v)

    path = tmp_path / "two.xlsx"
    wb.save(path)
    return str(path)


@pytest.fixture
def detect_source(two_layout_xlsx):
    return {
        "id": "two", "type": "google_sheet", "local_path": two_layout_xlsx,
        "sheet_pattern": r"^\d{2}\. ",
        "layout": {
            "detect_columns": True, "header_row": 3,
            "stop_markers": ["CHANGE REQUEST (CR)/FEEDBACK USER"],
            "valid_status": ["Done", "Progress", "Hold", "Back Log", "To Do"],
        },
    }


def test_both_layouts_parsed(detect_source):
    recs = GoogleSheetAdapter(detect_source).fetch()
    by_issue = {r.issue: r for r in recs}
    assert "Pengecekan akses VM" in by_issue       # 2025 layout
    assert "Scan Accunetix" in by_issue            # 2026 layout (NO kosong)


def test_2025_columns_mapped(detect_source):
    r = {x.issue: x for x in GoogleSheetAdapter(detect_source).fetch()}["Pengecekan akses VM"]
    assert r.pic == "Rio"
    assert r.status == "Done"
    assert r.keterangan == "ok"
    assert r.date_for("tgl_request").isoformat() == "2025-05-06"


def test_2026_columns_mapped(detect_source):
    r = {x.issue: x for x in GoogleSheetAdapter(detect_source).fetch()}["Scan Accunetix"]
    assert r.no is None
    assert r.pic == "Rio"
    assert r.status == "Progress"
    assert r.priority == "High"
    assert r.aplikasi == "Aplikasi CAT Sertifikasi Operator Radio"


def test_recap_noise_dropped(detect_source):
    recs = GoogleSheetAdapter(detect_source).fetch()
    # Baris rekap "CAT" dgn status "Low" harus tidak ikut.
    assert not any(r.issue == "CAT" for r in recs)
    assert all(r.status in {"Done", "Progress", "Hold", "Back Log", "To Do", "Tanpa Status"} for r in recs)
