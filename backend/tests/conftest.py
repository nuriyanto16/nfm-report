"""Fixture: bangun xlsx sintetis yang meniru layout sheet bulanan asli."""
from __future__ import annotations

from datetime import datetime

import openpyxl
import pytest


@pytest.fixture(autouse=True)
def _isolate_auth_store(tmp_path, monkeypatch):
    """Setiap test pakai file store user/role terpisah (jangan sentuh repo)."""
    monkeypatch.setenv("AUTH_STORE", str(tmp_path / "auth.json"))


def _write_month(ws, banner: str):
    # Baris 1: judul, baris 2 kosong, baris 3 header, baris 4 banner bulan.
    ws["A1"] = "CHANGE REQUEST (CR)/FEEDBACK USER"
    headers = ["NO", "ISSUE", "PATH MENU", "TANGGAL REQUEST/ISSUE",
               "TANGGAL MULAI", "TANGGAL ESTIMASI", "TANGGAL SELESAI",
               "REQUEST BY", "PIC", "KATEGORI", "STATUS", "PRIORITY",
               "STATUS DEPLOY", "Last Update", "", "KETERANGAN"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=3, column=i, value=h)
    ws["A4"] = banner


def _row(ws, r, no, issue, treq, status, pic, prio="High", katg="Fix Bug/Error"):
    ws.cell(row=r, column=1, value=no)
    ws.cell(row=r, column=2, value=issue)
    if treq:
        ws.cell(row=r, column=4, value=treq)
    ws.cell(row=r, column=9, value=pic)
    ws.cell(row=r, column=10, value=katg)
    ws.cell(row=r, column=11, value=status)
    ws.cell(row=r, column=12, value=prio)


@pytest.fixture
def synthetic_xlsx(tmp_path):
    wb = openpyxl.Workbook()
    # Sheet non-bulanan yang harus diabaikan.
    wb.active.title = "Summary"
    wb.active["A1"] = "ringkasan"

    ws = wb.create_sheet("06. Juni 2026")
    _write_month(ws, "BULAN JUNI 2026")
    ws["A5"] = "1. Pengujian"
    _row(ws, 6, 1.0, "Bug notifikasi WA", datetime(2026, 6, 3, 12, 30), "Done", "Hasan")
    _row(ws, 7, 2.0, "Tambah fitur export", datetime(2026, 6, 18, 9, 30), "Progress", "Rijal")
    _row(ws, 8, 3.0, "Cek log email", None, "Back Log", "Hasan")  # tgl_request kosong
    _row(ws, 9, 4.0, "Issue tanpa status", datetime(2026, 6, 20, 10, 0), None, "Ardia")  # status kosong
    # Noise: NO ada tapi ISSUE kosong -> harus di-skip.
    _row(ws, 10, 5.0, None, datetime(2026, 6, 21), "Done", "X")
    # Noise: header berulang.
    ws.cell(row=11, column=1, value="STATUS")
    ws["A12"] = "2. Kalibrasi"
    _row(ws, 13, 1.0, "Kalibrasi fitur copy", datetime(2026, 6, 16, 12, 0), "Done", "Mu'in")

    # --- blok legacy non-SIMS: harus dihentikan parsing-nya ---
    ws["A20"] = "CHANGE REQUEST (CR)/FEEDBACK USER"
    for i, h in enumerate(["NO", "ISSUE"], start=1):
        ws.cell(row=22, column=i, value=h)
    ws["A23"] = "BULAN DESEMBER 2025"
    ws["A24"] = "1. Aplikasi CAT Sertifikasi Operator Radio"
    _row(ws, 25, 1.0, "JANGAN ikut terhitung (legacy)", datetime(2025, 12, 1), "Done", "Z")

    # Bulan kedua untuk menguji multi-tab.
    ws2 = wb.create_sheet("05. Mei 2026")
    _write_month(ws2, "BULAN MEI 2026")
    ws2["A5"] = "1. Pengujian"
    _row(ws2, 6, 1.0, "Task Mei", datetime(2026, 5, 10, 9, 0), "Done", "Hasan")

    path = tmp_path / "synthetic.xlsx"
    wb.save(path)
    return str(path)


@pytest.fixture
def synthetic_source(synthetic_xlsx):
    return {
        "id": "test-src",
        "name": "Test Source",
        "type": "google_sheet",
        "local_path": synthetic_xlsx,
        "sheet_pattern": r"^\d{2}\. .+ \d{4}$",
        "layout": {
            "header_row": 3,
            "data_start_row": 4,
            "columns": {
                "no": 0, "issue": 1, "path_menu": 2, "tgl_request": 3,
                "tgl_mulai": 4, "tgl_estimasi": 5, "tgl_selesai": 6,
                "request_by": 7, "pic": 8, "kategori": 9, "status": 10,
                "priority": 11, "status_deploy": 12, "last_update": 13,
                "keterangan": 15,
            },
            "section_markers": {
                "1. Pengujian": "Pengujian (SIMPEL)",
                "2. Kalibrasi": "Kalibrasi",
                "3. Mobile": "Mobile",
                "4. Infra": "Infra",
            },
            "stop_markers": [
                "CHANGE REQUEST (CR)/FEEDBACK USER",
                "BULAN DESEMBER 2025",
            ],
            "valid_status": ["Done", "Progress", "Hold", "Back Log", "To Do"],
        },
    }
